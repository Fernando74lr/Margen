from .data.secret.credentials import server, username, password
from django.http.response import HttpResponse
from openpyxl.utils import get_column_letter
from .functions.tools import format_date, thin_border, enterprises, get_time, format_date_prefered
from openpyxl.styles import Alignment, Font, PatternFill
from django.shortcuts import render
from .models import SqlServerConn
from openpyxl import Workbook
import pyodbc
from sqlalchemy import create_engine
import pymssql

query_cache = []
beginDateG = ''
endDateG = ''

def table(request):
    return render(request, 'report/table_filter.html')

def clean_cache():
    global query_cache
    query_cache = [] # Clean cache
    print('\nCLEANING CACHE\n')

def reports(request):
    clean_cache()
    return render(request, 'report/reports.html')

def get_enterprises_names():
    database = 'CompacWAdmin'

    conn = pyodbc.connect('DRIVER={SQL Server};'+
                        'SERVER=' + server + ';'+
                        'DATABASE=' + database + ';'+
                        'UID=' + username + ';'+
                        'PWD=' + password + ';')

    cursor = conn.cursor()
    cursor.execute("SELECT CNOMBREEMPRESA, CRUTADATOS FROM Empresas")
    enterprisesDB = cursor.fetchall()

    enterprises_names = []

    for enterprise in enterprisesDB:
        enterprises_names.append([
            enterprise[0],
            enterprise[1].split('\\').pop()
        ])

    enterprises_names.pop(0)
    return enterprises_names

def get_available_databases():
    engine = create_engine(f'mssql+pymssql://{username}:{password}@{server}', deprecate_large_types=True)
    conn2 = engine.connect()
    databases = conn2.execute("select name FROM sys.databases;")
    available_databases = []

    for db in databases:
        if 'adCOM' in db['name']:
            # print(db["name"])
            available_databases.append(db["name"])
    
    return available_databases

def get_clients(request, beginDate, endDate):
    global beginDateG
    global endDateG

    enterprisesX = get_enterprises_names()

    print('\n')
    print(enterprisesX[0])
    print(enterprisesX[1])
    print('\n')

    beginDateG = beginDate.split('-')
    endDateG = endDate.split('-')

    beginDate = list(beginDate.split("/"))
    endDate = list(endDate.split("/"))

    print(f"\n{beginDateG} - {endDateG}\n")
    print(f"\n{beginDate} - {endDate}\n")

    global query_cache
    result = []

    available_databases = get_available_databases()
    # not_available_databases = []

    if len(query_cache) == 0:
        print("\nGETTING DATA\n")

        for database in enterprisesX:
            if database[1] in available_databases:
                try:
                    conn = pyodbc.connect('DRIVER={SQL Server};'+
                                        'SERVER=' + server + ';'+
                                        'DATABASE=' + database[1] + ';'+
                                        'UID=' + username + ';'+
                                        'PWD=' + password + ';')
                    cursor = conn.cursor()
                
                    cursor.execute("SELECT "+
                                "CSERIEDOCUMENTO, CFOLIO, CFECHA, "+
                                "CIDCLIENTEPROVEEDOR, CRAZONSOCIAL, CRFC, "+
                                "CCANCELADO, "+
                                "CNETO, CIMPUESTO1, CTOTAL, "+
                                "CMETODOPAG, CGUIDDOCUMENTO, CUSUARIO, CIDDOCUMENTO "+
                                "FROM admDocumentos " +
                                f"WHERE CFECHA BETWEEN '{''.join(beginDate)}' AND '{''.join(endDate)}'")
                    
                    # Expected input: 20200623 | 20200624

                    dataDocumentos = cursor.fetchall()
                    temp = []
                    
                    for query in dataDocumentos:            
                        query = list(query)
                        query.insert(0, database[1]) # Database name
                        query.insert(0, database[0]) # Enterprise name
                        observations = cursor.execute(f"SELECT COBSERVAMOV FROM admMovimientos WHERE CIDDOCUMENTO='{query[len(query)-1]}'")

                        maxLen = 0
                        observation_final = ''

                        # Get the longest observation
                        for observ in observations:
                            if str(observ[0]) != 'None':
                                if len(str(observ[0])) > maxLen:
                                    maxLen = len(observ[0])
                                    observation_final = str(observ[0])

                        query.append(observation_final)
                        cursor.fetchall()
                        temp.append(query)
                    result.append(temp)
                except NameError:
                    print("\nAn exception occurred:\n" )
                    print(NameError)
            else:
                print(f"\n'{database[1]}' database is not available")
                # not_available_databases.append(database[1])
        
        query_cache = result
    else:
        print("\nUSING CACHE\n")
        result = query_cache
    
    # if len(result) == len(enterprisesX):
    #     print("\nALL RIGHT\n")
    # else:
    #     print("SOMETHING WENT WRONG")

    return render(request, 'report/reports.html', {'SqlServerConn' : result})

def clients_report(request):
    global query_cache
    # OPEN WORKBOOK AND HEADER DETAILS
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws['A1'] = f'REPORTE DE CLIENTES - {format_date()} - {get_time()}'
    ws['E1'] = f'RANGO: {beginDateG[2]}/{beginDateG[1]}/{beginDateG[0]} - {endDateG[2]}/{endDateG[1]}/{endDateG[0]}'
    ws.merge_cells('A1:D1')
    ws.merge_cells('E1:F1')

    # HEADERS
    ws['A2'] = 'SERIE Y FOLIO'
    ws['B2'] = 'CORPORACIÓN'
    ws['C2'] = 'FECHA'
    ws['D2'] = 'RFC'
    ws['E2'] = 'ID CLIENTE PROVEEDOR'
    ws['F2'] = 'RAZÓN SOCIAL'
    ws['G2'] = 'CANCELADO'
    ws['H2'] = 'NETO'
    ws['I2'] = 'IMPUESTO 1'
    ws['J2'] = 'TOTAL'
    ws['K2'] = 'MÉTODO DE PAGO'
    ws['L2'] = 'GUID DOCUMENTO'
    ws['M2'] = 'USUARIO'
    ws['N2'] = 'ID DOCUMENTO'
    ws['O2'] = 'OBSERV. MOVIM.'

    # FILTERS
    FullRange = "A2:" + get_column_letter(ws.max_column) + str(ws.max_row)
    ws.auto_filter.ref = FullRange

    last_column = 0

    for enterprise in query_cache:
        for data in enterprise:
            if len(data[16]) > last_column:
                last_column = len(data[16])

    # ALIGNMENTS, COLORS AND DIMENSIONS
    dimensions = [22.14, 23.57, 13.14, 15.14, 34.57, 24.71, 20.43, 12, 19.71, 13, 28.57, 40.43, 16.29, 25.43, last_column]
   
    ws.row_dimensions[1].height = 26.25
    ws.row_dimensions[2].height = 42.75

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].font = Font(size="16", color="FF0000", b=True)

    ws['E1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E1'].font = Font(size="16", color="FF0000", b=True)

    for col in range(15):
        ws.cell(row=2, column=col+1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col+1).fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type = "solid")
        ws.cell(row=2, column=col+1).font = Font(size="16", color="FFFFFF")
        ws.cell(row=2, column=col+1).border = thin_border

    # Column size
    for i, column_width in enumerate(dimensions):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 1

    counter = 3

    # CREATE EXCEL
    if len(query_cache) > 0:
        print("\nCREATING EXCEL\n")
        for enterprise in query_cache:
            for data in enterprise:
                # Serie Documento
                ws.cell(row=counter, column=1).value = f'{data[2]}{int(data[3])}'
                ws.cell(row=counter, column=1).font = Font(size="12")
                ws.cell(row=counter, column=1).border = thin_border
                # Corporación
                ws.cell(row=counter, column=2).value = data[0]
                ws.cell(row=counter, column=2).font = Font(size="12")
                ws.cell(row=counter, column=2).border = thin_border
                # Fecha
                ws.cell(row=counter, column=3).value = format_date_prefered(data[4])
                ws.cell(row=counter, column=3).font = Font(size="12")
                ws.cell(row=counter, column=3).border = thin_border
                # RFC
                ws.cell(row=counter, column=4).value = data[7]
                ws.cell(row=counter, column=4).font = Font(size="12")
                ws.cell(row=counter, column=4).border = thin_border
                # ID Cliente Proveedor
                ws.cell(row=counter, column=5).value = str(data[5]).split(' ')[0]
                ws.cell(row=counter, column=5).font = Font(size="12")
                ws.cell(row=counter, column=5).border = thin_border
                # Razón Social
                ws.cell(row=counter, column=6).value = data[6]
                ws.cell(row=counter, column=6).font = Font(size="12")
                ws.cell(row=counter, column=6).border = thin_border
                # Cancelado
                ws.cell(row=counter, column=7).value = 'CANCELADO' if data[8] else ' - '
                ws.cell(row=counter, column=7).font = Font(size="12")
                ws.cell(row=counter, column=7).border = thin_border
                # Neto
                ws.cell(row=counter, column=8).value = data[9]
                ws.cell(row=counter, column=8).font = Font(size="12")
                ws.cell(row=counter, column=8).border = thin_border
                # Impuesto1
                ws.cell(row=counter, column=9).value = data[10]
                ws.cell(row=counter, column=9).font = Font(size="12")
                ws.cell(row=counter, column=9).border = thin_border
                # Total
                ws.cell(row=counter, column=10).value = data[11]
                ws.cell(row=counter, column=10).font = Font(size="12")
                ws.cell(row=counter, column=10).border = thin_border
                # Método de pago
                ws.cell(row=counter, column=11).value = str(data[12]).split(' ')[0]
                ws.cell(row=counter, column=11).font = Font(size="12")
                ws.cell(row=counter, column=11).border = thin_border
                # GUID Documento
                ws.cell(row=counter, column=12).value = data[13]
                ws.cell(row=counter, column=12).font = Font(size="12")
                ws.cell(row=counter, column=12).border = thin_border
                # Usuario
                ws.cell(row=counter, column=13).value = data[14]
                ws.cell(row=counter, column=13).font = Font(size="12")
                ws.cell(row=counter, column=13).border = thin_border
                # ID Documento
                ws.cell(row=counter, column=14).value = str(data[15]).split(' ')[0]
                ws.cell(row=counter, column=14).font = Font(size="12")
                ws.cell(row=counter, column=14).border = thin_border
                # Observ. Movim.
                ws.cell(row=counter, column=15).value = data[16] if data[16] else ' - '
                ws.cell(row=counter, column=15).font = Font(size="12")
                ws.cell(row=counter, column=15).border = thin_border

                counter+=1
    else:
        print("\nCACHE EMPTY\n")

    # FILE DETAILS AND FORMAT
    filename = 'Reporte_Clientes.xlsx'
    response = HttpResponse(content_type = 'application/ms-excel')
    content = 'attachment; filename = {0}'.format(filename)
    response['Content-Disposition'] = content
    wb.save(response)
    clean_cache()

    return response